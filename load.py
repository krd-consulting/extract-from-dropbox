import enum
import openpyxl
import os
from pathlib import Path
import pandas
import sqlalchemy
import sys
from dotenv import load_dotenv
from typing import List
from sqlalchemy import select
from sqlalchemy.orm import Mapped
from sqlalchemy.orm import mapped_column
from sqlalchemy.orm import relationship

class DocumentType(enum.Enum):
    BUDGET = 'BUDGET'
    CONTRACT = 'CONTRACT'
    FINANCIAL_REPORT = 'FINANCIAL-REPORT'

class FileName():
    def __init__(self, filename):
        self.filename = filename
        self.set_components()

    def set_components(self):
       self.filename_components = Path(self.filename).stem.split('_')
       self.document_type = DocumentType(self.filename_components[0])
       self.agency_code = self.filename_components[1]
       self.program_code = self.filename_components[2]
       self.set_fiscal_year()

    def set_fiscal_year(self):
        startYear = f"20{int(self.filename_components[3]) - 1}"
        endYear = self.filename_components[3]
        self.year = f"{startYear}/{endYear}"

class FinancialReportFileName(FileName):
    def set_components(self):
       super().set_components()
       self.set_quarter()
    
    def set_quarter(self):
        if self.filename_components[4] == 'Q1':
            self.quarter = 1

        if self.filename_components[4] == 'Q2':
            self.quarter = 2

        if self.filename_components[4] == 'Q3':
            self.quarter = 3
        
        if self.filename_components[4] == 'Q4':
            self.quarter = 4



class Base(sqlalchemy.orm.DeclarativeBase):
    pass

class Program(Base):
    __tablename__ = 'Programs'

    ID: Mapped[int] = mapped_column(primary_key=True)
    CODE: Mapped[str] = mapped_column(sqlalchemy.String(32), nullable=False)
    Name: Mapped[str] = mapped_column(sqlalchemy.String(64), nullable=False)

    contracts: Mapped[List['Contract']] = relationship(back_populates='program') 

class Contract(Base):
    __tablename__ = 'Contracts'

    ID: Mapped[int] = mapped_column(primary_key=True)
    TotalContract: Mapped[int] = mapped_column(sqlalchemy.Numeric(10, 2))
    ProgramID: Mapped[int] = mapped_column(sqlalchemy.ForeignKey('Programs.ID'), nullable=False)
    FiscalYear: Mapped[int] = mapped_column(sqlalchemy.NVARCHAR(8), nullable=False)

    program: Mapped['Program'] = relationship(back_populates='contracts')

class BudgetItem(Base):
    __tablename__ = 'BudgetItems'

    ID: Mapped[int] = mapped_column(primary_key=True)
    Name: Mapped[str] = mapped_column(sqlalchemy.VARCHAR(200), nullable=False)
    Amount: Mapped[int] = mapped_column(sqlalchemy.Numeric(10, 2))
    Category: Mapped[str] = mapped_column(sqlalchemy.VARCHAR(200))
    ContractID: Mapped[int] = mapped_column(sqlalchemy.INTEGER, nullable=False)
    FromFileName: Mapped[str] = mapped_column(sqlalchemy.VARCHAR(200), nullable=False)

    expense_allocations: Mapped[List['ExpenseAllocation']] = relationship(
            back_populates='budget_item',
            primaryjoin="and_(BudgetItem.ID == ExpenseAllocation.BudgetItemID)"
        )

class ExpenseAllocation(Base):
    __tablename__ = 'ExpenseAllocations'

    ID: Mapped[int] = mapped_column(primary_key=True)
    Q1AllocatedAmount: Mapped[int] = mapped_column(sqlalchemy.Numeric(10, 2))
    Q2AllocatedAmount: Mapped[int] = mapped_column(sqlalchemy.Numeric(10, 2))
    Q3AllocatedAmount: Mapped[int] = mapped_column(sqlalchemy.Numeric(10, 2))
    Q4AllocatedAmount: Mapped[int] = mapped_column(sqlalchemy.Numeric(10, 2))
    BudgetItemID: Mapped[int] = mapped_column(sqlalchemy.ForeignKey('BudgetItems.ID'), nullable=False)
    FromFileName: Mapped[str] = mapped_column(sqlalchemy.VARCHAR(200), nullable=False)
    Quarter: Mapped[int] = mapped_column(sqlalchemy.Integer(), nullable=False)

    budget_item: Mapped['BudgetItem'] = relationship(back_populates='expense_allocations')

def populate_budget_items_table(session, core_session, items, filename : FileName):
    """
    Using a filename, populate table with budget items from a contract
    (i.e. given program and year).

    Before inserting items, existing items from the same contract
    are deleted.
    """

    contract = get_contract(core_session, filename.program_code, filename.year)

    try:
        contract_id = contract.ID
    except sqlalchemy.exc.NoResultFound:
        # TODO: log
        # WorkFolder
        # FileName
        # Operation
        # Source
        return
    except sqlalchemy.exc.MultipleResultsFound:
        # TODO: log
        # WorkFolder
        # FileName
        # Operation
        # Source
        return

    delete_statement = (
        sqlalchemy.delete(BudgetItem)
        .where(BudgetItem.ContractID == contract_id)
    )

    session.execute(delete_statement)
    

    for i, item in enumerate(items):
        items[i]['ContractID'] = contract_id

    # TODO: log on insert
    # Optimize: bulk insert
    session.execute(
        sqlalchemy.insert(BudgetItem),
        items
    )

def get_contract(session, program_code, year):
    statement = (
        select(Contract)
        .join(Contract.program)
        .where(
            sqlalchemy.and_(Program.CODE == program_code, Contract.FiscalYear == year)
        )
    )

    # TODO: catch and log when No Result Found
    contract = session.scalars(statement).one()

    return contract

def populate_contracts_table():
    pass

def read_new_expense_allocations():
    pass

def read_new_budget_items(path):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    START = (2, 4)

    bold_cells_count = 0
    items = []
    for row in sheet.iter_rows(min_row=START[1], min_col=START[0], max_col=6):        
        item_name_cell = row[0]
        item_amount_cell = row[1]

        if(item_name_cell.font.b is True):
            # we always start at a bolded row
            bold_cells_count += 1
            current_item_category = item_name_cell.value
        
            # we're at a subtotal row
            if(bold_cells_count % 2 == 0):
                current_item_category = ''
                continue
        
        else:
            if(item_name_cell.value is None):
                # Blank line item or in between categories
                continue
        
            # insert expense allocation item
            items.append({
                'Name': item_name_cell.value,
                'FromFileName': path,
                'Category': current_item_category,
                'Amount': int(item_amount_cell.value or 0)
            })
        
    return items

def read_new_contracts():
    pass
    
def load_new_files(extracted_folder):
    pass

def read_financial_report(filename):
    path = filename

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

    reporting_sheet = workbook['Reporting']

    START = (2, 4)

    bold_cells_count = 0
    items = []
    for row in reporting_sheet.iter_rows(min_row=START[1], min_col=START[0], max_col=6):
        allocation_name_cell = row[0]
        allocation_values_columns_range = row[1:5]


        if(allocation_name_cell.font.b is True):
            # we always start at a bolded row
            bold_cells_count += 1
            current_expense_allocation_category = allocation_name_cell.value
        
            # we're at a subtotal row
            if(bold_cells_count % 2 == 0):
                current_expense_allocation_category = ''
                continue
        
        else:
            if(allocation_name_cell.value is None):
                # Blank line item or in between categories
                continue

            # insert expense allocation item
            items.append({
                'FromFileName': path,
                'Name': allocation_name_cell.value,
                'Category': current_expense_allocation_category,
                'Q1AllocatedAmount': str(allocation_values_columns_range[0].value or 0),
                'Q2AllocatedAmount': str(allocation_values_columns_range[1].value or 0),
                'Q3AllocatedAmount': str(allocation_values_columns_range[2].value or 0),
                'Q4AllocatedAmount': str(allocation_values_columns_range[3].value or 0)
            })
    return items

def populate_expense_allocations_table(session, core_session, items, filename : FinancialReportFileName):
    program_code = filename.program_code
    contract = get_contract(core_session, program_code, filename.year)
    
    # Delete all old expense allocation items for a contract and quarter.
    delete_statement = (
        sqlalchemy.delete(ExpenseAllocation).where(
            sqlalchemy.and_(
                BudgetItem.ID == ExpenseAllocation.BudgetItemID,
                BudgetItem.ContractID == contract.ID,
                ExpenseAllocation.Quarter == filename.quarter
            )
        )
    )
    session.execute(delete_statement)

    new_expense_allocations = pandas.DataFrame(items)
    new_expense_allocations['Quarter'] = filename.quarter
    budget_items_statement = (
        select(BudgetItem)
        .where(BudgetItem.ContractID == contract.ID)
        
    )
    budget_items = pandas.read_sql(budget_items_statement, session.bind)
    budget_items = budget_items.rename(columns={"ID": "BudgetItemID", "FromFileName": "FromBudgetFileName"})

    # Add `Category` on the join clause to make sure expense allocations with the same name but different
    # category get associated with the right budget items.
    all_expense_allocations = new_expense_allocations.merge(budget_items, on=['Name', 'Category'], how='left')
    with_matching_budget_item = new_expense_allocations.merge(budget_items, on=['Name', 'Category'], how='inner')
    # TODO: Log expense allocations with no matched budget items
    no_matched_budget_items = all_expense_allocations[all_expense_allocations['BudgetItemID'].isnull()]
    print(no_matched_budget_items)

    new_items = with_matching_budget_item[['BudgetItemID', 'Q1AllocatedAmount', 'Q2AllocatedAmount', 'Q3AllocatedAmount', 'Q4AllocatedAmount', 'FromFileName', 'Quarter']]
    print(new_items)
    new_items.to_sql(name='ExpenseAllocations', con=session.connection(), if_exists='append', index=False)

if __name__ == '__main__':
    load_dotenv()

    SERVER = os.environ['DATABASE_SERVER']
    CORE_DATABASE = os.environ['CORE_DATABASE']
    DATABASE = os.environ['DATABASE']
    CORE_ENGINE = sqlalchemy.create_engine('mssql+pyodbc://' + SERVER + '/' + CORE_DATABASE + '?TrustServerCertificate=yes&trusted_connection=yes&driver=ODBC+Driver+18+for+SQL+Server')
    ENGINE = sqlalchemy.create_engine('mssql+pyodbc://' + SERVER + '/' + DATABASE + '?TrustServerCertificate=yes&trusted_connection=yes&driver=ODBC+Driver+18+for+SQL+Server')

    # TODO: test manually loading in files
    path_to_file = sys.argv[1]
    # TODO: try to validate file name and throw error for invalid ones.
    filename = FileName(path_to_file)
    
    if(filename.document_type == DocumentType.BUDGET):
        budget_items = read_new_budget_items(path_to_file)
        with sqlalchemy.orm.Session(ENGINE) as session:
            populate_budget_items_table(session, sqlalchemy.orm.Session(CORE_ENGINE), budget_items, filename)
            session.commit()
    elif(filename.document_type == DocumentType.FINANCIAL_REPORT):
        with sqlalchemy.orm.Session(ENGINE) as session:
            report_filename = FinancialReportFileName(filename.filename)
            expense_allocation_items = read_financial_report(report_filename.filename)
            populate_expense_allocations_table(session, sqlalchemy.orm.Session(CORE_ENGINE), expense_allocation_items, report_filename)
            session.commit()